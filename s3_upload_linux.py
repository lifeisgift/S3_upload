# -*- coding: utf-8 -*-
# PROJECT LONGBOW - AMAZON S3 UPLOAD TOOL WITH BREAK-POINT RESUMING
import os
import sys
import json
import base64
import boto3
from boto3.session import Session
from botocore.client import Config
from concurrent import futures
from configparser import ConfigParser, RawConfigParser, NoOptionError
import time
import datetime
import hashlib
import logging
from pathlib import PurePosixPath, Path
import platform
import codecs
import shutil
import zipfile
import math
import threading
import openpyxl
from openpyxl import load_workbook
from multiprocessing import Process #直接使用进程方式
from multiprocessing import Pool  #使用进程池方式
from os.path import join, getsize
os.system("")  # workaround for some windows system to print color

global JobType, SrcFileIndex, DesProfileName, DesBucket, S3Prefix, MaxRetry, MaxThread, SrcFileAggreRatio, \
    MaxParallelFile, StorageClass, ifVerifyMD5, DontAskMeToClean, LoggingLevel, \
    SrcDir, SrcBucket, SrcProfileName,LocalTempPath, AggreFileList, ifDeleteTempZipfile, LocalFolderSize, \
    ifRecoverFromLast, LastZipIndexFile

# 读取配置参数
def set_config():
    sys_para = sys.argv
    file_path = os.path.split(sys_para[0])[0]
    gui = False

    JobType_list = ['LOCAL_TO_S3', 'S3_TO_S3', 'ALIOSS_TO_S3']
    StorageClass_list = ['STANDARD', 'REDUCED_REDUNDANCY', 'STANDARD_IA', 'ONEZONE_IA', 'INTELLIGENT_TIERING',
                         'GLACIER', 'DEEP_ARCHIVE']
    config_file = os.path.join(file_path, 's3_upload_config.ini')

    # If no config file, read the default config
    if not os.path.exists(config_file):
        config_file += '.default'
        print("No customized config, use the default config")
    cfg = ConfigParser()
    print(f'Reading config file: {config_file}')

    # Get local config value
    try:
        global JobType, SrcFileIndex, DesProfileName, DesBucket, S3Prefix, MaxRetry, MaxThread, SrcFileAggreRatio, \
            MaxParallelFile, StorageClass, ifVerifyMD5, DontAskMeToClean, LoggingLevel, \
            SrcDir, SrcBucket, SrcProfileName,LocalTempPath, ifDeleteTempZipfile, LocalFolderSize, \
            ifRecoverFromLast, LastZipIndexFile
        cfg.read(config_file, encoding='utf-8-sig')
        JobType = cfg.get('Basic', 'JobType')
        SrcFileIndex = cfg.get('Basic', 'SrcFileIndex')
        DesProfileName = cfg.get('Basic', 'DesProfileName')
        DesBucket = cfg.get('Basic', 'DesBucket')
        S3Prefix = cfg.get('Basic', 'S3Prefix')
        SrcFileAggreRatio = cfg.getint('Basic', 'SrcFileAggreRatio')
        LocalTempPath = cfg.get('Basic','LocalTempPath')
        Megabytes = 1024 * 1024
        ChunkSize = cfg.getint('Advanced', 'ChunkSize') * Megabytes
        MaxRetry = cfg.getint('Advanced', 'MaxRetry')
        MaxThread = cfg.getint('Advanced', 'MaxThread')
        MaxParallelFile = cfg.getint('Advanced', 'MaxParallelFile')
        StorageClass = cfg.get('Advanced', 'StorageClass')
        ifVerifyMD5 = cfg.getboolean('Advanced', 'ifVerifyMD5')
        ifDeleteTempZipfile = cfg.getboolean('Advanced', 'ifDeleteTempZipfile')
        DontAskMeToClean = cfg.getboolean('Advanced', 'DontAskMeToClean')
        LoggingLevel = cfg.get('Advanced', 'LoggingLevel')
        LocalFolderSize = cfg.getint('Basic','LocalFolderSize')
        ifRecoverFromLast = cfg.getboolean('Basic','ifRecoverFromLast')
        LastZipIndexFile = cfg.get('Basic','LastZipIndexFile')

        try:
            SrcDir = cfg.get('LOCAL_TO_S3', 'SrcDir')
        except NoOptionError:
            SrcDir = ''

    except Exception as e:
        print("ERR loading s3_upload_config.ini", str(e))
        input('PRESS ENTER TO QUIT')
        sys.exit(0)


    S3Prefix = str(PurePosixPath(S3Prefix))  # 去掉结尾的'/'，如果有的话
    if S3Prefix == '/' or S3Prefix == '.':
        S3Prefix = ''
    # 校验
    if JobType not in JobType_list:
        print(f'ERR JobType: {JobType}, check config file: {config_file}')
        input('PRESS ENTER TO QUIT')
        sys.exit(0)
    # Finish set_config()
    return ChunkSize

# 创建log文件
def set_log():
    logger = logging.getLogger()
    # File logging
    if not os.path.exists("./log"):
        os.system("mkdir log")
    this_file_name = os.path.splitext(os.path.basename(__file__))[0]
    file_time = datetime.datetime.now().isoformat().replace(':', '-')[:19]
    log_file_name = './log/' + this_file_name + '-' + file_time + '.log'
    print('Logging to file:', os.path.abspath(log_file_name))
    print('Logging level:', LoggingLevel)
    fileHandler = logging.FileHandler(filename=log_file_name, encoding='utf-8')
    fileHandler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s - %(message)s'))
    logger.addHandler(fileHandler)
    # Screen stream logging
    streamHandler = logging.StreamHandler()
    streamHandler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s - %(message)s'))
    logger.addHandler(streamHandler)
    # Loggin Level
    logger.setLevel(logging.WARNING)
    if LoggingLevel == 'INFO':
        logger.setLevel(logging.INFO)
    elif LoggingLevel == 'DEBUG':
        logger.setLevel(logging.DEBUG)
    return logger, log_file_name

# 聚合压缩图片
def zip_file(src_dir, local_temp_path, SrcFileAggreRatio):
    global AggreFileList
    tempname = src_dir.replace("/","_")
    j = 0
    nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')

    AggreFileList =local_temp_path + "/" + nowTime + "-list.txt" #每次生成的唯一的压缩文件list
    AggreFileMappingSheet = local_temp_path + "/" + nowTime + "-list.xlsx" #每次生成原始文件和聚合文件的mapping 表格
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    file = open(AggreFileList,"a")
    for dirpath, dirnames, filenames in os.walk(src_dir):
        sheet.append({'A':str(dirpath)})  #原始文件名写入A列
        #sheet.append({'B':str(dirnames)})  #原始路径名写入B列
        i = 0
        fpath = dirpath.replace(src_dir,'')
        fpath = fpath and fpath + os.sep or ''

        #当前目录中的文件数目
        filenumberinfolder = len(filenames)
        #print("文件数目:",filenumberinfolder)
        for filename in filenames:
            #判断是否是第0个文件，创建新的文件名
            if i % SrcFileAggreRatio == 0 :
                zip_name = local_temp_path + "/" + tempname + fpath.replace("/", "_") + str(j) + ".zip"
                logger.info(f'Create aggregate file: {zip_name}')
                sheet.append({'B': zip_name})   #聚合文件名写入D列

                file.write(zip_name)
                file.write('\r\n')
                z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
                j = j + 1

            z.write(os.path.join(dirpath, filename), fpath + filename)
            #print ('==压缩成功==')
            sheet.append({'C': str(filename)})  # 原始文件名写入C列

            #判断是否到达压缩比设置或者文件夹最后一个文件，如果是，则关闭文件句柄
            if i % SrcFileAggreRatio == SrcFileAggreRatio - 1 or i + 1 == filenumberinfolder:
                z.close()

            i = i + 1

        #凑不到压缩比整数的时候，关闭文件句柄，此时需要添加判断文件是否存在？
        #z.close()

        #每个目录结束后，j清零，下一个目录生成的文件名从头开始
        j = 0
    #z.close()
    file.close()
    workbook.save(AggreFileMappingSheet)
    logger.info(f'File Mapping is saved to : {AggreFileMappingSheet}')

# Get object list on S3
def get_s3_file_list(*, s3_client, bucket, S3Prefix, no_prefix=False):
    logger.info('Get s3 file list ' + bucket)

    # For delete prefix in des_prefix
    if S3Prefix == '':
        # 目的bucket没有设置 Prefix
        dp_len = 0
    else:
        # 目的bucket的 "prefix/"长度
        dp_len = len(S3Prefix) + 1

    paginator = s3_client.get_paginator('list_objects_v2')
    __des_file_list = []
    try:
        response_iterator = paginator.paginate(
            Bucket=bucket,
            Prefix=S3Prefix
        )
        for page in response_iterator:
            if "Contents" in page:
                for n in page["Contents"]:
                    key = n["Key"]
                    if no_prefix:
                        key = key[dp_len:]
                    __des_file_list.append({
                        "Key": key,
                        "Size": n["Size"]
                    })
        logger.info(f'Bucket list length：{str(len(__des_file_list))}')
    except Exception as err:
        logger.error(str(err))
        input('PRESS ENTER TO QUIT')
        sys.exit(0)
    return __des_file_list


# Check single file on S3
def head_s3_single_file(s3_client, bucket):
    try:
        response_fileList = s3_client.head_object(
            Bucket=bucket,
            Key=str(Path(S3Prefix)/SrcFileIndex)
        )
        file = [{
            "Key": str(Path(S3Prefix)/SrcFileIndex),
            "Size": response_fileList["ContentLength"]
        }]
    except Exception as err:
        logger.error(str(err))
        input('PRESS ENTER TO QUIT')
        sys.exit(0)
    return file


# Get all exist object list on S3
def get_uploaded_list(s3_client):
    logger.info('Get unfinished multipart upload')
    NextKeyMarker = ''
    IsTruncated = True
    __multipart_uploaded_list = []
    while IsTruncated:
        list_multipart_uploads = s3_client.list_multipart_uploads(
            Bucket=DesBucket,
            Prefix=S3Prefix,
            MaxUploads=1000,
            KeyMarker=NextKeyMarker
        )
        IsTruncated = list_multipart_uploads["IsTruncated"]
        NextKeyMarker = list_multipart_uploads["NextKeyMarker"]
        if NextKeyMarker != '':
            for i in list_multipart_uploads["Uploads"]:
                __multipart_uploaded_list.append({
                    "Key": i["Key"],
                    "Initiated": i["Initiated"],
                    "UploadId": i["UploadId"]
                })
                logger.info(f'Unfinished upload, Key: {i["Key"]}, Time: {i["Initiated"]}')
    return __multipart_uploaded_list


# Jump to handle next file
class NextFile(Exception):
    pass

def uploadThread_small(srcfile, prefix_and_key):
    print(f'\033[0;32;1m--->Uploading\033[0m {srcfile["Key"]} - small file')
    with open(os.path.join(SrcDir, srcfile["Key"]), 'rb') as data:
        for retryTime in range(MaxRetry + 1):
            try:
                pstart_time = time.time()
                chunkdata = data.read()
                chunkdata_md5 = hashlib.md5(chunkdata)
                s3_dest_client.put_object(
                    Body=chunkdata,
                    Bucket=DesBucket,
                    Key=prefix_and_key,
                    ContentMD5=base64.b64encode(chunkdata_md5.digest()).decode('utf-8'),
                    StorageClass=StorageClass
                )
                pload_time = time.time() - pstart_time
                pload_bytes = len(chunkdata)
                pload_speed = size_to_str(int(pload_bytes / pload_time)) + "/s"
                print(f'\033[0;34;1m    --->Complete\033[0m {srcfile["Key"]} - small file - {pload_speed}')
                break
            except Exception as e:
                logger.warning(f'Upload small file Fail: {srcfile["Key"]}, '
                               f'{str(e)}, Attempts: {retryTime}')
                if retryTime >= MaxRetry:
                    logger.error(f'Fail MaxRetry Download/Upload small file: {srcfile["Key"]}')
                    return "MaxRetry"
                else:
                    time.sleep(5 * retryTime)
    return


def upload_file(*, srcfile, desFilelist, UploadIdList, ChunkSize_default):  # UploadIdList就是multipart_uploaded_list
    logger.info(f'Start file: {srcfile["Key"]}')
    prefix_and_key = srcfile["Key"]
    if JobType == 'LOCAL_TO_S3':
        prefix_and_key = str(PurePosixPath(S3Prefix) / srcfile["Key"])
    if srcfile['Size'] >= ChunkSize_default:
        try:
            # 循环重试3次（如果MD5计算的ETag不一致）
            for md5_retry in range(3):
                # 检查文件是否已存在，存在不继续、不存在且没UploadID要新建、不存在但有UploadID得到返回的UploadID
                response_check_upload = check_file_exist(srcfile=srcfile,
                                                         desFilelist=desFilelist,
                                                         UploadIdList=UploadIdList)
                if response_check_upload == 'UPLOAD':
                    logger.info(f'New upload: {srcfile["Key"]}')
                    response_new_upload = s3_dest_client.create_multipart_upload(
                        Bucket=DesBucket,
                        Key=prefix_and_key,
                        StorageClass=StorageClass
                    )
                    # logger.info("UploadId: "+response_new_upload["UploadId"])
                    reponse_uploadId = response_new_upload["UploadId"]
                    partnumberList = []
                elif response_check_upload == 'NEXT':
                    logger.info(f'Duplicated. {srcfile["Key"]} same size, goto next file.')
                    raise NextFile()
                else:
                    reponse_uploadId = response_check_upload

                    # 获取已上传partnumberList
                    partnumberList = checkPartnumberList(srcfile, reponse_uploadId)

                # 获取索引列表，例如[0, 10, 20]
                response_indexList, ChunkSize_auto = split(srcfile, ChunkSize_default)

                # 执行分片upload
                upload_etag_full = uploadPart(uploadId=reponse_uploadId,
                                              indexList=response_indexList,
                                              partnumberList=partnumberList,
                                              srcfile=srcfile,
                                              ChunkSize_auto=ChunkSize_auto)

                # 合并S3上的文件
                response_complete = completeUpload(reponse_uploadId=reponse_uploadId,
                                                   srcfileKey=srcfile["Key"],
                                                   len_indexList=len(response_indexList))
                logger.info(f'FINISH: {srcfile["Key"]} TO {response_complete["Location"]}')

                # 检查文件MD5
                if ifVerifyMD5:
                    if response_complete["ETag"] == upload_etag_full:
                        logger.info(f'MD5 ETag Matched - {srcfile["Key"]} - {response_complete["ETag"]}')
                        break
                    else:  # ETag 不匹配，删除S3的文件，重试
                        logger.warning(f'MD5 ETag NOT MATCHED {srcfile["Key"]}( Destination / Origin ): '
                                       f'{response_complete["ETag"]} - {upload_etag_full}')
                        s3_dest_client.delete_object(
                            Bucket=DesBucket,
                            Key=prefix_and_key
                        )
                        UploadIdList = []
                        logger.warning('Deleted and retry upload {srcfile["Key"]}')
                    if md5_retry == 2:
                        logger.warning('MD5 ETag NOT MATCHED Exceed Max Retries - {srcfile["Key"]}')
                else:
                    break
        except NextFile:
            pass

    # Small file procedure
    else:
        # Check file exist
        for f in desFilelist:
            if f["Key"] == prefix_and_key and \
                    (srcfile["Size"] == f["Size"]):
                logger.info(f'Duplicated. {prefix_and_key} same size, goto next file.')
                return
        # 找不到文件，或文件Size不一致 Submit upload
        if JobType == 'LOCAL_TO_S3':
            uploadThread_small(srcfile, prefix_and_key)
    return


# Compare file exist on desination bucket
def check_file_exist(*, srcfile, desFilelist, UploadIdList):
    # 检查源文件是否在目标文件夹中
    prefix_and_key = srcfile["Key"]
    if JobType == 'LOCAL_TO_S3':
        prefix_and_key = str(PurePosixPath(S3Prefix) / srcfile["Key"])
    for f in desFilelist:
        if f["Key"] == prefix_and_key and \
                (srcfile["Size"] == f["Size"]):
            return 'NEXT'  # 文件完全相同
    # 找不到文件，或文件不一致，要重新传的
    # 查Key是否有未完成的UploadID
    keyIDList = []
    for u in UploadIdList:
        if u["Key"] == prefix_and_key:
            keyIDList.append(u)
    # 如果找不到上传过的Upload，则从头开始传
    if not keyIDList:
        return 'UPLOAD'
    # 对同一个Key（文件）的不同Upload找出时间最晚的值
    UploadID_latest = keyIDList[0]
    for u in keyIDList:
        if u["Initiated"] > UploadID_latest["Initiated"]:
            UploadID_latest = u
    return UploadID_latest["UploadId"]


# Check parts number exist on S3
def checkPartnumberList(srcfile, uploadId):
    try:
        prefix_and_key = srcfile["Key"]
        if JobType == 'LOCAL_TO_S3':
            prefix_and_key = str(PurePosixPath(S3Prefix) / srcfile["Key"])
        partnumberList = []
        PartNumberMarker = 0
        IsTruncated = True
        while IsTruncated:
            response_uploadedList = s3_dest_client.list_parts(
                Bucket=DesBucket,
                Key=prefix_and_key,
                UploadId=uploadId,
                MaxParts=1000,
                PartNumberMarker=PartNumberMarker
            )
            NextPartNumberMarker = response_uploadedList['NextPartNumberMarker']
            IsTruncated = response_uploadedList['IsTruncated']
            if NextPartNumberMarker > 0:
                for partnumberObject in response_uploadedList["Parts"]:
                    partnumberList.append(partnumberObject["PartNumber"])
            PartNumberMarker = NextPartNumberMarker
        if partnumberList:  # 如果为0则表示没有查到已上传的Part
            logger.info("Found uploaded partnumber: " + json.dumps(partnumberList))
    except Exception as checkPartnumberList_err:
        logger.error("checkPartnumberList_err" + json.dumps(checkPartnumberList_err))
        input('PRESS ENTER TO QUIT')
        sys.exit(0)
    return partnumberList

# split the file into a virtual part list of index, each index is the start point of the file
def split(srcfile, ChunkSize):
    partnumber = 1
    indexList = [0]
    if int(srcfile["Size"] / ChunkSize) + 1 > 10000:
        ChunkSize = int(srcfile["Size"] / 10000) + 1024  # 对于大于10000分片的大文件，自动调整Chunksize
        logger.info(f'Size excess 10000 parts limit. Auto change ChunkSize to {ChunkSize}')

    while ChunkSize * partnumber < srcfile["Size"]:  # 如果刚好是"="，则无需再分下一part，所以这里不能用"<="
        indexList.append(ChunkSize * partnumber)
        partnumber += 1
    return indexList, ChunkSize

# upload parts in the list
def uploadPart(*, uploadId, indexList, partnumberList, srcfile, ChunkSize_auto):
    partnumber = 1  # 当前循环要上传的Partnumber
    total = len(indexList)
    md5list = [hashlib.md5(b'')] * total
    complete_list = []
    # 线程池Start
    with futures.ThreadPoolExecutor(max_workers=MaxThread) as pool:
        for partStartIndex in indexList:
            # start to upload part
            if partnumber not in partnumberList:
                dryrun = False
            else:
                dryrun = True
            # upload 1 part/thread, or dryrun to only caculate md5
            if JobType == 'LOCAL_TO_S3':
                pool.submit(uploadThread,
                            uploadId=uploadId,
                            partnumber=partnumber,
                            partStartIndex=partStartIndex,
                            srcfileKey=srcfile["Key"],
                            total=total,
                            md5list=md5list,
                            dryrun=dryrun,
                            complete_list=complete_list,
                            ChunkSize=ChunkSize_auto)
            partnumber += 1
    # 线程池End
    logger.info(f'All parts uploaded - {srcfile["Key"]} - size: {srcfile["Size"]}')

    # Local upload 的时候考虑传输过程中文件会变更的情况，重新扫描本地文件的MD5，而不是用之前读取的body去生成的md5list
    if ifVerifyMD5 and JobType == 'LOCAL_TO_S3':
        md5list = cal_md5list(indexList=indexList,
                              srcfileKey=srcfile["Key"],
                              ChunkSize=ChunkSize_auto)
    # 计算所有分片列表的总etag: cal_etag
    digests = b"".join(m.digest() for m in md5list)
    md5full = hashlib.md5(digests)
    cal_etag = '"%s-%s"' % (md5full.hexdigest(), len(md5list))
    return cal_etag


# convert bytes to human readable string
def size_to_str(size):
    def loop(integer, remainder, level):
        if integer >= 1024:
            remainder = integer % 1024
            integer //= 1024
            level += 1
            return loop(integer, remainder, level)
        else:
            return integer, round(remainder / 1024, 1), level

    units = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
    integer, remainder, level = loop(int(size), 0, 0)
    if level+1 > len(units):
        level = -1
    return f'{integer+remainder} {units[level]}'


# 本地文件重新计算一次MD5
def cal_md5list(*, indexList, srcfileKey, ChunkSize):
    logger.info(f'Re-read local file to calculate MD5 again: {srcfileKey}')
    md5list = []
    with open(os.path.join(SrcDir, srcfileKey), 'rb') as data:
        for partStartIndex in indexList:
            data.seek(partStartIndex)
            chunkdata = data.read(ChunkSize)
            chunkdata_md5 = hashlib.md5(chunkdata)
            md5list.append(chunkdata_md5)
    return md5list

# Single Thread Upload one part, from local to s3
def uploadThread(*, uploadId, partnumber, partStartIndex, srcfileKey, total, md5list, dryrun, complete_list, ChunkSize):
    prefix_and_key = str(PurePosixPath(S3Prefix) / srcfileKey)
    if not dryrun:
        print(f'\033[0;32;1m--->Uploading\033[0m {srcfileKey} - {partnumber}/{total}')
    pstart_time = time.time()
    with open(os.path.join(SrcDir, srcfileKey), 'rb') as data:
        retryTime = 0
        while retryTime <= MaxRetry:
            try:
                data.seek(partStartIndex)
                chunkdata = data.read(ChunkSize)
                chunkdata_md5 = hashlib.md5(chunkdata)
                md5list[partnumber - 1] = chunkdata_md5
                if not dryrun:
                    s3_dest_client.upload_part(
                        Body=chunkdata,
                        Bucket=DesBucket,
                        Key=prefix_and_key,
                        PartNumber=partnumber,
                        UploadId=uploadId,
                        ContentMD5=base64.b64encode(chunkdata_md5.digest()).decode('utf-8')
                    )
                    # 这里对单个part上传做了 MD5 校验，后面多part合并的时候会再做一次整个文件的
                break
            except Exception as err:
                retryTime += 1
                logger.info(f'UploadThreadFunc log: {srcfileKey} - {str(err)}')
                logger.info(f'Upload Fail - {srcfileKey} - Retry part - {partnumber} - Attempt - {retryTime}')
                if retryTime > MaxRetry:
                    logger.error(f'Quit for Max retries: {retryTime}')
                    input('PRESS ENTER TO QUIT')
                    sys.exit(0)
                time.sleep(5 * retryTime)  # 递增延迟重试
    complete_list.append(partnumber)
    pload_time = time.time() - pstart_time
    pload_bytes = len(chunkdata)
    pload_speed = size_to_str(int(pload_bytes / pload_time)) + "/s"
    if not dryrun:
        print(f'\033[0;34;1m    --->Complete\033[0m {srcfileKey} '
              f'- {partnumber}/{total} \033[0;34;1m{len(complete_list) / total:.2%} - {pload_speed}\033[0m')
    return


# Complete multipart upload, get uploadedListParts from S3 and construct completeStructJSON
def completeUpload(*, reponse_uploadId, srcfileKey, len_indexList):
    # 查询S3的所有Part列表uploadedListParts构建completeStructJSON
    prefix_and_key = srcfileKey
    if JobType == 'LOCAL_TO_S3':
        prefix_and_key = str(PurePosixPath(S3Prefix) / srcfileKey)
    uploadedListPartsClean = []
    PartNumberMarker = 0
    IsTruncated = True
    while IsTruncated:
        response_uploadedList = s3_dest_client.list_parts(
            Bucket=DesBucket,
            Key=prefix_and_key,
            UploadId=reponse_uploadId,
            MaxParts=1000,
            PartNumberMarker=PartNumberMarker
        )
        NextPartNumberMarker = response_uploadedList['NextPartNumberMarker']
        IsTruncated = response_uploadedList['IsTruncated']
        if NextPartNumberMarker > 0:
            for partObject in response_uploadedList["Parts"]:
                ETag = partObject["ETag"]
                PartNumber = partObject["PartNumber"]
                addup = {
                    "ETag": ETag,
                    "PartNumber": PartNumber
                }
                uploadedListPartsClean.append(addup)
        PartNumberMarker = NextPartNumberMarker
    if len(uploadedListPartsClean) != len_indexList:
        logger.warning(f'Uploaded parts size not match - {srcfileKey}')
        input('PRESS ENTER TO QUIT')
        sys.exit(0)
    completeStructJSON = {"Parts": uploadedListPartsClean}

    # S3合并multipart upload任务
    response_complete = s3_dest_client.complete_multipart_upload(
        Bucket=DesBucket,
        Key=prefix_and_key,
        UploadId=reponse_uploadId,
        MultipartUpload=completeStructJSON
    )
    logger.info(f'Complete merge file {srcfileKey}')
    return response_complete


# Compare local file list and s3 list
def compare_local_to_s3(src_file_list):
    logger.info('Comparing destination and source ...')
    fileList = src_file_list
    desFilelist = get_s3_file_list(s3_client=s3_dest_client,
                                   bucket=DesBucket,
                                   S3Prefix=S3Prefix,
                                   no_prefix=True)
    deltaList = []
    for source_file in fileList:
        if source_file not in desFilelist:
            deltaList.append(source_file)
    if not deltaList:
        logger.warning('All source files are in destination Bucket/Prefix. Job well done.')
    else:
        logger.warning(f'There are {len(deltaList)} files not in destination or not the same size. List:')
        for delta_file in deltaList:
            logger.warning(str(delta_file))
    return


# get the filelist and size
# #获得文件名和大小的列表, 用于和当前S3的对象列表进行比较
def local_file_list_withsize():
    __src_file_list = []
    global AggreFileList

    for line in open(AggreFileList):
        line = line.strip()
        file_size = os.path.getsize(line)
        line = line[line.rfind('/'):]
        line = line.strip('/')
        __src_file_list.append({
            "Key": line,
            "Size": file_size
        })
    return __src_file_list

#upload with threading
#利用多进程上传，提高上传效率
def upload_with_thread(localkey, prefix_and_key):
    global MaxThread,LocalTempPath,DesBucket,StorageClass

    #S3和本地文件比对
    response = s3_dest_client.list_objects_v2(Bucket=DesBucket, Prefix=prefix_and_key, MaxKeys=1)
    file_exists = response['KeyCount'] > 0
    if file_exists :
        #file from S3
        response = s3_dest_client.head_object(Bucket=DesBucket,Key=prefix_and_key)
        objectsize = response['ContentLength']
        #file from local
        absfilename = str(PurePosixPath(LocalTempPath) / localkey)
        local_file_size = os.path.getsize(absfilename)
        #check if with the same size
        if local_file_size == objectsize :
            logger.info(f'Object {localkey} exist in S3 bucket and with the same size, will pass it!')
        else :
            logger.info(f'Object {localkey} exist in S3 bucket and with different size, will replace it!')
            with open(os.path.join(LocalTempPath, localkey), 'rb') as data:
                filedata = data.read()
            # upload
            s3_dest_client.put_object(
                Body=filedata,
                Bucket=DesBucket,
                StorageClass=StorageClass,
                Key=prefix_and_key
            )
            logger.info(f'Upload newer aggregate file: {localkey}')
    else :
        # if no file with the same name, upload directly
        with open(os.path.join(LocalTempPath, localkey), 'rb') as data:
            filedata = data.read()
        # upload
        s3_dest_client.put_object(
            Body=filedata,
            Bucket=DesBucket,
            StorageClass=StorageClass,
            Key=prefix_and_key
        )
        logger.info(f'Upload aggregate file: {localkey}')

    # delete temp file
    if ifDeleteTempZipfile == True :
        absfilename = str(PurePosixPath(LocalTempPath) / localkey)
        if os.path.exists(absfilename):
            time.sleep(1)
            os.remove(absfilename)
            logger.info(f'Remove temp aggregate file: {absfilename}')
            #多线程并发模式下利用返回值及下面的threadres函数获取进程返回值
    message = "Upload aggregate file:" + localkey
    return message

#此callback函数用于获取进程返回值
def threadres(arg):
    return arg

# get the total size of a folder
# 添加锁，解决文件冲突问题，若在计算过程中，某个文件刚好被删除
def getdirsize():
    global LocalTempPath
    size = 0
    for root, dirs, files in os.walk(LocalTempPath):
        for f in files:
            # 获取锁。
            lock.acquire()
            if os.path.exists(f):
                fp = os.path.join(root, f)
                size += os.path.getsize(fp)
            # 释放锁，
            lock.release()
        #size += sum([getsize(join(root, name)) for name in files])
    return size

if __name__ == '__main__':
    start_time = datetime.datetime.now()
    ChunkSize_default = set_config()
    logger, log_file_name = set_log()

    # 生成锁对象，全局唯一
    lock = threading.Lock()
    # 生成锁，用于计算文件夹大小的时候防止文件被删除了，获取文件大小出错

    #进程pool方式的返回值
    res_list = []

    # Define s3 client
    s3_config = Config(max_pool_connections=200)
    s3_dest_client = Session(profile_name=DesProfileName).client('s3', config=s3_config)
    # Check destination S3 writable
    try:
        logger.info(f'Checking write permission for: {DesBucket}')
        s3_dest_client.put_object(
            Bucket=DesBucket,
            Key=str(PurePosixPath(S3Prefix) / 'access_test'),
            Body='access_test_content'
        )
    except Exception as e:
        logger.error(f'Can not write to {DesBucket}/{S3Prefix}, {str(e)}')
        input('PRESS ENTER TO QUIT')
        sys.exit(0)

    # aggregate file under each folder and write to a temp folder
    logger.info('Aggregate file in every folder to defined ratio and upload immediately')

    #创建线程池，不超过设置值
    pool = Pool(MaxThread)

    #打包压缩文件，并同时上传，打包完成后即给线程池进行上传
    str_from = f'{SrcDir}'

    #先判断是恢复上传还是全新上传，全新上传逻辑不变，恢复上传逻辑新写-10.11
    if ifRecoverFromLast == False :
        logger.info('This is a fresh upload!')

        global AggreFileList

        tempname = SrcDir.replace("/", "_")
        j = 0
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')

        AggreFileList = LocalTempPath + "/" + nowTime + "-list.txt"  # 每次生成的唯一的压缩文件list
        AggreFileMappingSheet = LocalTempPath + "/" + nowTime + "-list.xlsx"  # 每次生成原始文件和聚合文件的mapping 表格
        workbook = openpyxl.Workbook()
        sheet = workbook.active #获取excel表格的当前活动页

        file = open(AggreFileList, "a")
        for dirpath, dirnames, filenames in os.walk(SrcDir):
            sheet.append({'A': str(dirpath)})  # 原始文件名写入A列
            # sheet.append({'B':str(dirnames)})  #原始路径名写入B列
            i = 0
            fpath = dirpath.replace(SrcDir, '')
            fpath = fpath and fpath + os.sep or ''

            # 当前目录中的文件数目
            filenumberinfolder = len(filenames)
            # print("文件数目:",filenumberinfolder)
            for filename in filenames:
                try:

                    # 判断是否是第0个文件，创建新的文件名
                    if i % SrcFileAggreRatio == 0:
                        key = tempname + fpath.replace("/", "_") + str(j) + ".zip"
                        zip_name = LocalTempPath + "/" + tempname + fpath.replace("/", "_") + str(j) + ".zip"
                        logger.info(f'Create aggregate file: {zip_name}')
                        sheet.append({'B': zip_name})  # 聚合文件名写入D列

                        file.write(zip_name)
                        file.write('\r\n')
                        z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
                        j = j + 1

                    z.write(os.path.join(dirpath, filename), fpath + filename)
                    # print ('==压缩成功==')
                    sheet.append({'C': str(filename)})  # 原始文件名写入C列
                except UnicodeEncodeError:
                    logger.error(f': This filename is illegal in unicode: {dirpath}/{filename}')

                # 判断是否到达压缩比设置或者文件夹最后一个文件，如果是，则关闭文件句柄
                if i % SrcFileAggreRatio == SrcFileAggreRatio - 1 or i + 1 == filenumberinfolder:
                    z.close()
                    #此时文件生成，可以进行上传
                    # 利用进程池方式并发上传
                    prefix_and_key = str(PurePosixPath(S3Prefix) / key)
                    res = pool.apply_async(func=upload_with_thread, args=(key, prefix_and_key,), callback=threadres)
                    res_list.append(res)


                i = i + 1

                #检测当前压缩目录下文件数目，若超过数量，则主进程sleep
                taobalFolderSize = getdirsize()
                taobalFolderSizeMB = taobalFolderSize/1024/1024/1024
                #print("current folder size is:",taobalFolderSizeMB)
                if taobalFolderSizeMB > LocalFolderSize/2 :
                    resleep = taobalFolderSizeMB/LocalFolderSize*10
                    print("current folder size is:", taobalFolderSizeMB)
                    time.sleep(2*resleep)

            # 每个目录结束后，j清零，下一个目录生成的文件名从头开始
            j = 0
        # z.close()
        file.close()
        workbook.save(AggreFileMappingSheet)
        logger.info(f'File Mapping is saved to : {AggreFileMappingSheet}')

    #若是恢复上传，判断原先上传位置，然后从最后一个压缩包看是重新打包上传
    elif ifRecoverFromLast == True :
        logger.info('This upload will recover from last breaking point!')
        #上次的excel记录文件,oldAggreFileList
        global oldAggreFileList
        oldAggreFileList = LocalTempPath + "/" + LastZipIndexFile
        #当前目录下的zip文件列表
        global oldZipFileListinFolder
        oldZipFileListinFolder = []
        #当前excel中的zip文件列表
        global oldZipFileListinExcel
        oldZipFileListinExcel = []
        #需要上传到额zip文件列表
        global uploadZipFileList
        uploadZipFileList = []
        #最后一个压缩文件的路径及文件名
        global lastFilePath           #excel中最后一个上传路径
        lastFilePath = ""
        global lastFileName             #最后一个压缩过的文件
        lastFileName = ""
        global pathListinExcel         #excel中的文件路径列表
        pathListinExcel = []
        global lastPathRow             #最后一个目录在excel中的行数
        global uploadedFileInLastPath   #最后一个目录下已经上传文件列表
        uploadedFileInLastPath = []

        oldExcel = load_workbook(oldAggreFileList)
        worksheet1 = oldExcel.active
        #step1, check if there is any zip packages still in the temp folder which should be uploaded last time.
        zipList = os.listdir(LocalTempPath)
        for cur_file in zipList:
            temp_path = os.path.join(LocalTempPath,cur_file )
            if os.path.isfile(temp_path): #if this is file
                if temp_path.endswith("zip"):
                    oldZipFileListinFolder.append(temp_path)
        #print(oldZipFileListinFolder)
        #excel中所有zip文件
        for cell in worksheet1['B']:
            if cell.value != None :
                oldZipFileListinExcel.append(cell.value)
        # print(oldZipFileListinExcel)

        #倒叙查找最后一个打包路径
        for cell in reversed(worksheet1['A']):
            if cell.value != None :
                lastFilePath = cell.value
                logger.info(f'Last path in Excel  is: {lastFilePath}')
                lastPathRow = cell.row
                break

        #倒叙查找最后一个打包文件
        for cell in reversed(worksheet1['C']):
            if cell.value != None:
                lastFileName = cell.value
                #print("Last file name is : ",lastFileName)
                logger.info(f'Last file in Excel  is: {lastFileName}')
                break
        #列出已经打包路径
        for cell in reversed(worksheet1['A']):
            if cell.value != None:
                pathListinExcel.append(cell.value)
        #列出最后一个路径下已经上传文件
        for cell in reversed(worksheet1['C']):
            if cell.row >= lastPathRow:
                break
            if cell.value != None:
                uploadedFileInLastPath.append(cell.value)

        #对比两个列表，文件若在excel中，说明打包完成，则需要上传。开始多进程上传流程
        uploadZipFileList = set(oldZipFileListinFolder).intersection(set(oldZipFileListinExcel))
        logger.info(f'Following zip file need to be check and upload : {uploadZipFileList}')
        #print(uploadZipFileList)
        for zipFIle in uploadZipFileList :
            path,key = os.path.split(zipFIle)
            #print(key)
            logger.info(f'Recover start, filename is:,{key}')
            prefix_and_key = str(PurePosixPath(S3Prefix) / key)
            res = pool.apply_async(func=upload_with_thread, args=(key, prefix_and_key,), callback=threadres)
            res_list.append(res)

        # step2, 从excel中上次位置开始继续打包上传流程
        logger.info('Existing zip file uploaded done, Start from last breaking point!')

        tempname = SrcDir.replace("/", "_")
        j = 0
        nowTime = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')

        AggreFileList = LocalTempPath + "/" + nowTime + "-list.txt"  # 每次生成的唯一的压缩文件list
        AggreFileMappingSheet = LocalTempPath + "/" + nowTime + "-list.xlsx"  # 每次生成原始文件和聚合文件的mapping 表格
        workbook = openpyxl.Workbook()
        sheet = workbook.active  # 获取excel表格的当前活动页

        file = open(AggreFileList, "a")
        for dirpath, dirnames, filenames in os.walk(SrcDir):
            i = 0
            fpath = dirpath.replace(SrcDir, '')
            fpath = fpath and fpath + os.sep or ''

            #从breaking 路径往后的文档都需要上传, 对于当前未完成的目录单独处理
            if dirpath == lastFilePath :
                sheet.append({'A': str(dirpath)}) #在A列写入路径
                filenumberinfolder = len(filenames)
                for filename in filenames:  #本循环是一个路径下所有文件的压缩和上传
                    if filename not in uploadedFileInLastPath: #若文件不在已经上传目录中，则执行压缩上传逻辑
                        try:
                            # 判断是否是第0个文件，创建新的文件名
                            if i % SrcFileAggreRatio == 0:
                                key = tempname + fpath.replace("/", "_") + nowTime + str(j) + ".zip"
                                zip_name = LocalTempPath + "/" + tempname + fpath.replace("/", "_") + nowTime + str(j) + ".zip"
                                logger.info(f'Create aggregate file: {zip_name}')
                                sheet.append({'B': zip_name})  # 聚合文件名写入D列

                                file.write(zip_name)
                                file.write('\r\n')
                                z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
                                j = j + 1

                            z.write(os.path.join(dirpath, filename), fpath + filename)
                            # print ('==压缩成功==')
                            sheet.append({'C': str(filename)})  # 原始文件名写入C列
                        except UnicodeEncodeError:
                            logger.error(f': This filename is illegal in unicode: {dirpath}/{filename}')

                        # 判断是否到达压缩比设置或者文件夹最后一个文件，如果是，则关闭文件句柄
                        if i % SrcFileAggreRatio == SrcFileAggreRatio - 1 or i + 1 == filenumberinfolder:
                            z.close()
                            # 此时文件生成，可以进行上传
                            # 利用进程池方式并发上传
                            prefix_and_key = str(PurePosixPath(S3Prefix) / key)
                            res = pool.apply_async(func=upload_with_thread, args=(key, prefix_and_key,),
                                                   callback=threadres)
                            res_list.append(res)

                        i = i + 1

                        # 检测当前压缩目录下文件数目，若超过数量，则主进程sleep，等待一段时间
                        taobalFolderSize = getdirsize()
                        taobalFolderSizeMB = taobalFolderSize / 1024 / 1024 / 1024
                        # print("current folder size is:",taobalFolderSizeMB)
                        if taobalFolderSizeMB > LocalFolderSize / 2:
                            resleep = taobalFolderSizeMB / LocalFolderSize * 10
                            print("current folder size is:", taobalFolderSizeMB)
                            time.sleep(2 * resleep)


            #其他后续目录走这个分支，上传所有文件
            if dirpath not in pathListinExcel:
                sheet.append({'A': str(dirpath)})  # 原始文件名写入A列
                filenumberinfolder = len(filenames)
                for filename in filenames:

                    try:
                        # 判断是否是第0个文件，创建新的文件名
                        if i % SrcFileAggreRatio == 0:
                            key = tempname + fpath.replace("/", "_") + str(j) + ".zip"
                            zip_name = LocalTempPath + "/" + tempname + fpath.replace("/", "_") + str(j) + ".zip"
                            logger.info(f'Create aggregate file: {zip_name}')
                            sheet.append({'B': zip_name})  # 聚合文件名写入D列

                            file.write(zip_name)
                            file.write('\r\n')
                            z = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
                            j = j + 1

                        z.write(os.path.join(dirpath, filename), fpath + filename)
                        # print ('==压缩成功==')
                        sheet.append({'C': str(filename)})  # 原始文件名写入C列
                    except UnicodeEncodeError:
                        logger.error(f': This filename is illegal in unicode: {dirpath}/{filename}')

                    # 判断是否到达压缩比设置或者文件夹最后一个文件，如果是，则关闭文件句柄
                    if i % SrcFileAggreRatio == SrcFileAggreRatio - 1 or i + 1 == filenumberinfolder:
                        z.close()
                        # 此时文件生成，可以进行上传
                        # 利用进程池方式并发上传
                        prefix_and_key = str(PurePosixPath(S3Prefix) / key)
                        res = pool.apply_async(func=upload_with_thread, args=(key, prefix_and_key,), callback=threadres)
                        res_list.append(res)

                    i = i + 1

                    # 检测当前压缩目录下文件数目，若超过数量，则主进程sleep
                    taobalFolderSize = getdirsize()
                    taobalFolderSizeMB = taobalFolderSize / 1024 / 1024 / 1024
                    # print("current folder size is:",taobalFolderSizeMB)
                    if taobalFolderSizeMB > LocalFolderSize / 2:
                        resleep = taobalFolderSizeMB / LocalFolderSize * 10
                        print("current folder size is:", taobalFolderSizeMB)
                        time.sleep(2 * resleep)

            # 每个目录结束后，j清零，下一个目录生成的文件名从头开始
            j = 0
        # z.close()
        file.close()
        workbook.save(AggreFileMappingSheet)
        logger.info(f'File Mapping is saved to : {AggreFileMappingSheet}')

    #关闭进程池并输出日志
    pool.close()
    pool.join()
    for res in res_list:
        print(res.get())

    time_str = str(datetime.datetime.now() - start_time)
    print(f'\033[0;34;1mMISSION ACCOMPLISHED - Time: {time_str} \033[0m - FROM: {str_from} TO {DesBucket}/{S3Prefix}')
    print('Logged to file:', os.path.abspath(log_file_name))
    input('PRESS ENTER TO QUIT')
